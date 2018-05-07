# -*- coding: utf-8 -*-
"""
@Created parse.py on 17-3-1 AM 4:49
@Author: Gaobo.Xiao
@Version: V1
@license: Intel-ICG
"""
import re
import xlrd
import redis
import operator
import openpyxl
from io import BytesIO
from collections import OrderedDict
import traceback
from django.db import transaction
from django.core.cache import cache

from apps.rest.validation.serializers import *
from apps.rest.ecr.serializers import *


class ValidationReportParse(object):
    """
    1: Validation excel format
    2: Check the data is valid
    3: save the excel file stream to file
    4:
    """
    SECTION_INDEX = 0
    SECTION_START_MARK = "section_start"
    SECTION_END_MARK = "section_end"
    # store in menu config
    SECTION_NAME_MAP = {
                        "execution_summary": {"guid": True, "alias": "execution_summary", "type": "horizontal",
                                              "dialect":{"fail": "failed", "block":"blocked", "pass": "passed", "passrate":"pass_rate"},
                                              "expect_keys": ["category", "passed", "failed", "blocked", "na", "nc", "total", "pass_rate", "weight", "weighted_pass_rate"],
                                              "serializer_cls": "TestSummarySerializer"},
                        "test_summary": {"guid": True, "alias": "test_summary", "type": "vertical",
                                         "dialect": {"highlights": "highlight", "low_light": "lowlight"},
                                         "expect_keys": ["note", "highlight", "lowlight", "comment"],
                                         "optional_keys": [], "serializer_cls": "TestCommentSerializer"},
                        "test_environment": {"guid": True, "alias": "test_environment", "type": "no_head_data",
                                             "expect_keys": ["env_key", "env_value"],
                                             "serializer_cls": "TestEnvironmentSerializer"},
                        "klocwork_scan_result": {"guid": True, "alias": "klocwork_scan_result", "type": "horizontal",
                                                 "dialect":{"project": "project_name"},
                                                 "expect_keys": ["project_name", "open_issues", "need_to_solve", "critical", "error", "banned_functions"], "serializer_cls": "KlocworkScanResultSerializer"},
                        "klocwork_server": {"guid": True, "alias": "klocwork_server", "type": "no_head_data",
                                            "dialect":{"kw_server": "server_name"}, "expect_keys": ["server_name", 'server_url'],
                                            "optional_keys": ["server_url"], "serializer_cls": "KlocworkServerSerializer"},
                        "protexip_scan_result": {"guid": True, "alias": "protexip_scan_result", "type": "horizontal",
                                                 "expect_keys": ["project_name", "code_matches", "license_conflicts"],
                                                 "serializer_cls": "ProtexipScanResultSerializer"},
                        "protexip_server": {"guid": True, "alias": "protexip_ip_server", "type": "no_head_data",
                                            "dialect":{"ip_server": "server_name"}, "expect_keys": ["server_name", "server_url"],
                                            "optional_keys": ["server_url"], "serializer_cls": "ProtexipServerSerializer"},
                        "customize_url": {"guid": True, "alias": "customize_url", "type": "horizontal",
                                                     "dialect":{"urltext": "url_text"},
                                                     "expect_keys": ["component", "url_text", "url"], "serializer_cls": "ExecutionSummaryAccessSerializer"},
                        "hsd_info": {"guid": True, "alias": "test_buglist", "type": "horizontal",
                                     "dialect": {"summary": "title", "type": "bug_type", "id": "bug_id"},
                                     "expect_keys":["bug_id", "bug_type"],
                                     "serializer_cls": "TestBugsListSerializer",
                                     #"injection_keys": ["ingredient"],
                                     },
                        "test_suite":{"guid": True, "alias": "test_suite", "type": "horizontal",
                                      "dialect": {"category_name": "first_category", "suite_name": "second_category",
                                                  "category": "first_category", "suite": "second_category"},
                                      "expect_keys": ["first_category", "second_category", "comment", "weight"],
                                      "injection_keys": {"weight_passrate": "0"},
                                      "serializer_cls": "TestSuiteSerializer"},
                        "weekly_case_detail": {"guid": True, "type": "horizontal", "alias": "test_case_detail",
                                               "dialect": {"test_case_id": "case_id", "name": "case_name", "type2": "case_type",
                                                           "test_result": "out_come", "test_type": "case_type"},
                                               "expect_keys": ["case_name", "case_type",  "out_come", "exe_type"],
                                               "optional_keys": ["first_category", "second_category", "hsd", "hsd_url"],
                                               "serializer_cls": "WeeklyTestDetailSerializer",
                                               "ignore_keys":["out_come"]},#如果这条记录该值为空，则忽略掉这条记录
                        "pit_case_detail": {"guid": True, "type": "horizontal",
                                                 "expect_keys": ["case_id", "case_name", "out_come", "exe_type"],
                                                 "dialect": {"test_case_id":"case_id", "name": "case_name",
                                                             "test_result": "out_come"},
                                                 "alias": "test_case_detail",
                                                 # "injection_keys": {"exe_type": "Auto"},
                                                 "serializer_cls": "PITTestDetailSerializer",
                                                 "ignore_keys": ["out_come"]},
                        "pit_lite_case_detail": {"guid": True, "type": "horizontal",
                                                 "expect_keys": ["case_id", "case_name", "out_come", "exe_type"],
                                                 "dialect": {"test_case_id":"case_id", "name": "case_name",
                                                             "test_result": "out_come"},
                                                 "alias": "test_case_detail",
                                                 # "injection_keys": {"exe_type": "Auto"},
                                                 "serializer_cls": "PITLiteTestDetailSerializer",
                                                 "ignore_keys": ["out_come"]},
                        "cit_case_detail": {"guid": True, "type": "horizontal",
                                                    "expect_keys": ["case_id", "case_name", "out_come", "exe_type"],
                                                    "dialect": {"test_case_id": "case_id", "name": "case_name",
                                                                "test_result": "out_come"},
                                                    "alias": "test_case_detail",
                                                    # "injection_keys": {"exe_type": "Auto"},
                                                    "serializer_cls": "CITTestDetailSerializer",
                                                    "ignore_keys": ["out_come"]},
                        }

    RULES_MAP = {
        "sum":[
            {"name": "weight", "range": {"max": 1.01, "min": 0.99}, "serializer_cls": "TestSuiteSerializer"}
        ],
        "buildTime": {"name": "env_key", "required_value": "Build Time", "serializer_cls": "TestEnvironmentSerializer"}
    }


    def __init__(self, *args, **kwargs):

        self.file_path = kwargs.get('file_path')
        self.file_stream = kwargs.get('file_stream')

        if self.file_path is None and self.file_stream is None:
            raise Exception('excel data not found!')
        else:
            self._work_book = None
            self._sheet_name_map = OrderedDict()

        if self.file_path:
            self.file_stream = open(self.file_path, 'r').read()

        if self.file_stream:
            self.file_stream = self.file_stream.read()

        self._work_book = xlrd.open_workbook(file_contents=self.file_stream)
        for index, sheet in enumerate(self._work_book.sheets()):
            if not sheet.visibility:
                self._sheet_name_map[self.trim(sheet.name)] = {"alias": sheet.name, "index": index, "sections": OrderedDict()}

    @property
    def work_book(self):
        return self._work_book

    @property
    def sheet_name_map(self):
        return self._sheet_name_map

    @staticmethod
    def restore(file_path=None, memory_key=None):
        if memory_key:
            conn = redis.StrictRedis(host="127.0.0.1", port=6379, db=0)
            return ValidationReportParse(**{"file_stream": conn.get(memory_key)})
        else:
            return ValidationReportParse(**{"file_stream": open(file_path, 'r').read()})

    @classmethod
    def trim(cls, value):
        return re.sub(r'[-:;"\',\\/\s\n]+', '_', str(value).lower())

    def register(self, sheet, section, start_index, end_index, typo):
        _SECTION_NAME_MAP = self.SECTION_NAME_MAP
        if typo == "case_detail":
            if str(self.sheet_name_map['guid']['type']) == str(TestGuid.TYPE_STATUS_WEEKLY):
                _SECTION_NAME_MAP = {'case_detail': _SECTION_NAME_MAP['weekly_case_detail']}
            elif str(self.sheet_name_map['guid']['type']) == str(TestGuid.TYPE_STATUS_PIT):
                _SECTION_NAME_MAP = {'case_detail': _SECTION_NAME_MAP['pit_case_detail']}
            elif str(self.sheet_name_map['guid']['type']) == str(TestGuid.TYPE_STATUS_PIT_LITE):
                _SECTION_NAME_MAP = {'case_detail': _SECTION_NAME_MAP['pit_lite_case_detail']}
            elif str(self.sheet_name_map['guid']['type']) == str(TestGuid.TYPE_STATUS_CIT):
                _SECTION_NAME_MAP = {'case_detail': _SECTION_NAME_MAP['cit_case_detail']}
            _SECTION_NAME_MAP['case_detail']['alias'] = section

        for s in _SECTION_NAME_MAP:
            if section == _SECTION_NAME_MAP[s].get("alias"):
                if _SECTION_NAME_MAP[s]["type"] == 'horizontal':
                    start_index += 2
                    end_index -= 1
                    key_list = [cell.value for cell in sheet.row(start_index - 1)]
                elif _SECTION_NAME_MAP[s]["type"] == 'vertical':
                    key_list = []
                    start_index += 1
                    # end_index -= 1
                    for row in range(start_index, end_index):
                        key_list.append(sheet.cell(row, self.SECTION_INDEX).value)
                else:  # no_head_data
                    key_list = _SECTION_NAME_MAP[s]["expect_keys"]
                    start_index += 1
                    end_index -= 1

                sheet_name = self.trim(sheet.name)
                # set data start_index & end index & key list
                self._sheet_name_map[sheet_name]["sections"][s] = {
                    "key_list": key_list,
                    "start_index": start_index,
                    "end_index": end_index,
                    "type": typo,
                }
                self._sheet_name_map[sheet_name]["sections"][s].update(
                    _SECTION_NAME_MAP[s]
                )
                break

    def save(self, temporary=False):
        # save to redis
        if temporary:
            conn = redis.StrictRedis(host="127.0.0.1", port=6379, db=0)
            conn.set('temporary_file', self.file_stream)
        else:
            with open('save_new.xlsx', 'w') as f:
                f.write(self.file_stream)

    def format(self, lookup_keys, expect_keys, optional_keys, dialect):

        # ordered_list = {}

        # format primary key
        _lookup_keys = {}
        for index, lookup in enumerate(lookup_keys):
            lookup = self.trim(lookup.strip())
            for d, k in dialect.items():
                if d == lookup:
                    _lookup_keys[index] = k
                    break
            else:
                if lookup:
                    _lookup_keys[index] = lookup
        miss_key = set(expect_keys) - set(_lookup_keys.values())
        if miss_key:
            raise ICGException('ICG000001', ','.join(miss_key))
        # else:
        #     for expect in expect_keys:
        #         ordered_list[expect] = _lookup_keys.index(expect)
        #
        # # format optional key
        # for optional in optional_keys:
        #     try:
        #         ordered_list[optional] = lookup_keys.index(optional)
        #     except ValueError:
        #         pass
        #
        # ordered_list = [x[0] for x in sorted(ordered_list.items(), key=operator.itemgetter(1))]
        return _lookup_keys


    def check(self, section_map):
        sheet = self.work_book.sheet_by_index(section_map['index'])

        for k, v in section_map['sections'].items():
            # format key list mark
            try:
                keys_list = self.format(v['key_list'], v['expect_keys'], v.get('optional_keys', []), v.get('dialect', {}))
                data = []
                if v['type'] == 'vertical':
                    raw = {}
                    for row in range(v['start_index'], v['end_index']):
                        for vertical_key in v['key_list']:
                            if self.trim(sheet.cell(rowx=row, colx=0).value) == self.trim(vertical_key):
                                raw[v['dialect'][self.trim(vertical_key)]
                                if v['dialect'].get(self.trim(vertical_key))
                                else self.trim(vertical_key)] = sheet.cell(rowx=row, colx=1).value
                    if v.get('guid') and self.sheet_name_map['guid'].get('id'):
                        raw['guid'] = self.sheet_name_map['guid']['id']
                    data.append(raw)
                else:
                    #检查case_detail中的顺序是否满足
                    #   --》 category
                    #     --》 suite
                    #       --》 case
                    category_index, case_id_index, first_category, second_category = 0, 0, 0, 0
                    front_param = ''
                    #初始化警告变量
                    warning = ''
                    if 'detail' in v['alias']:
                        for _index, _key in enumerate(keys_list.values()):
                            if _key == 'category_type':
                                category_index = _index
                            # if _key == 'exe_type':
                            #     exe_type_index = _index
                            if _key == 'case_name':
                                name_index = _index
                            if _key == 'case_id':
                                case_id_index = _index

                    for row in range(v['start_index'], v['end_index'] + 1):
                        excel_row = row + 1
                        #protexip_scan_reult and klocwork_scan_result will cause one error with excel can't load data
                        cols = [str(sheet.cell(rowx=row, colx=__k).value)
                                if k in ['protexip_scan_result', 'klocwork_scan_result'] else sheet.cell(rowx=row, colx=__k).value
                                for __k, __v in keys_list.items()]
                        #检查ignore字段是否为空
                        #检查case_detail中的顺序是否满足
                        #==================================================================================================#
                        if category_index:
                            if not front_param:
                                if self.trim(cols[category_index].strip()) == 'category':
                                    front_param = 'category'
                                    first_category = cols[name_index]
                                    continue
                                else:
                                    raise ICGException('ICG100005', 'Row %d The order of category_type is wrong!' % excel_row)
                            elif front_param == 'category':
                                if self.trim(cols[category_index].strip()) == 'suite':
                                    front_param = 'suite'
                                    second_category = cols[name_index]
                                    continue
                                else:
                                    raise ICGException('ICG100005', 'Row %d The order of category_type is wrong!' % excel_row)
                            elif front_param == 'suite' or front_param == 'case':
                                if self.trim(cols[category_index].strip()) == 'case':
                                    front_param = 'case'
                                elif self.trim(cols[category_index].strip()) == 'suite':
                                    front_param = 'suite'
                                    second_category = cols[name_index]
                                    continue
                                elif self.trim(cols[category_index].strip()) == 'category' and front_param == 'case':
                                    front_param = 'category'
                                    first_category = cols[name_index]
                                    continue
                                elif self.trim(cols[category_index].strip()) == 'category' and front_param == 'suite':
                                    front_param = 'category'
                                    first_category = cols[name_index]
                                    #raise ICGException()  make warning!
                                    continue
                                else:
                                    raise ICGException('ICG100005', 'Row %d The order of category_type is wrong!' % excel_row)
                            else:
                                raise ICGException('ICG100005', 'Row %d The category_type is wrong!' % excel_row)
                        #==================================================================================================#
                        raw = dict(zip(keys_list.values(), cols))
                        _raw_keys = raw.keys()
                        #将指定为空字段忽略
                        _flag = False
                        for ignore in v.get("ignore_keys", []):
                            if ignore in v.get("expect_keys", []):
                                if not raw[ignore]:
                                    _flag = True
                        if _flag:
                            continue
                        #更新字典(现在Excel里已经有了first category 和 second category)
                        if v['alias'].endswith('casedetail_nocategory'):
                            if 'case_type' not in _raw_keys:
                                raise ICGException('ICG100004', 'Can not find [test type] in excel data')
                            if cols[case_id_index]:
                                # if not raw.get('first_category') and not raw.get('second_category'):
                                    first_category, second_category = self.get_contour_items_all_parent(
                                        self.sheet_name_map['guid']['project_id'],
                                        cols[case_id_index])
                                    if first_category and second_category:
                                        raw.update({'first_category': first_category,
                                                    'second_category': second_category})
                                    else:
                                        raise ICGException('ICG100004',
                                                           "Row %d Can't find match case from case management!" % excel_row)

                        if v['alias'].endswith('case_detail'):
                            if 'category_type' not in _raw_keys:
                                raise ICGException('ICG100004', 'Can not find [category type] in excel data')
                            if first_category and second_category:
                                raw.update({'first_category': first_category,
                                        'second_category': second_category})
                            else:
                                raise ICGException('ICG100004', "Row %d Can't find category from excel!"% excel_row)
                        #若存在hsd_id, 更新hsd_url
                        if 'issue_id' in _raw_keys:
                            hsd_url = self.update_hsd_url(raw['issue_id'])
                            raw.update({'hsd': raw['issue_id'],
                                        'hsd_url': hsd_url})
                        #pass_rate倍率调整
                        if 'pass_rate' in _raw_keys:
                            raw.update({'pass_rate': str(int(raw['pass_rate']*100))})
                        #更新HSD INFO中的BUG_url
                        if 'bug_id' in _raw_keys:
                            raw.update({'bug_id': self.update_bug_id(raw['bug_id']),
                                        'bug_url': self.update_hsd_url(raw['bug_id'])})
                        #==================================================================================================#
                        if v.get('guid') and self.sheet_name_map['guid'].get('id'):
                            raw['guid'] = self.sheet_name_map['guid']['id']
                        #注入缺少的字段
                        for _k, _v in v.get("injection_keys", {}).items():
                            raw[_k] = _v

                        data.append(raw)

                serializer = eval(v['serializer_cls'])(data=data, many=True)
                if self.sheet_name_map['guid']['execute_mode'] == 'new':
                    if serializer.is_valid():
                        msg, flag = self.check_validated_data(serializer.validated_data, self.RULES_MAP, v['serializer_cls'])
                        if flag:
                            v['errors'] = msg
                        v['validated_data'] = serializer.data
                    else:
                        v['errors'] = serializer.errors
                elif self.sheet_name_map['guid']['execute_mode'] == 'merge':
                    # msg, flag = self.check_validated_data(data, self.RULES_MAP, v['serializer_cls'])
                    # if flag:
                    #     v['errors'] = msg
                    v['validated_data'] = serializer.initial_data

            except ICGException as e:
                v['errors'] = e.error_detail
            except Exception as e:
                v['errors'] = str(e)

    def parse(self, sheet_name):
        sheet = self.get_sheet_by_alias(sheet_name)
        project_flag = section = msg = None
        sections = OrderedDict()

        # get section start&end index
        for index, section_mark in enumerate(sheet.col_values(colx=self.SECTION_INDEX)):
            if section_mark and self.trim(section_mark) == self.SECTION_START_MARK:
                # section mark followed section name
                section = str(sheet.cell(index, self.SECTION_INDEX+1).value).strip()
                if self.trim(section).endswith('casedetail_nocategory'):
                    if not self.sheet_name_map['guid']['project_id']:
                        section = None
                        project_flag = True
                        continue
                sections.setdefault(self.trim(section), {})
                # mark data start index
                sections[self.trim(section)]["start_index"] = index
                if self.trim(section).endswith('case_detail'):
                    sections[self.trim(section)]["owner"] = "case_detail"
                elif self.trim(section).endswith('casedetail_nocategory'):
                    sections[self.trim(section)]["owner"] = "case_detail"
                else:
                    sections[self.trim(section)]["owner"] = "extra"

            if section_mark and self.trim(section_mark) == self.SECTION_END_MARK:
                if section == str(sheet.cell(index, self.SECTION_INDEX+1).value).strip() and section :
                    sections.setdefault(self.trim(section), {})
                    # mark data end index
                    sections[self.trim(section)]["end_index"] = index
                else:
                    if not project_flag:
                        msg = str(sheet.cell(index, self.SECTION_INDEX+1).value).strip() + " section start flag not found"

        for k, v in sections.items():
            if v.get("start_index") is None:
                msg = k + " section start flag not found"
            if v.get("end_index") is None:
                msg = k + " section end flag not found"
            if msg:
                break
            else:
                self.register(sheet, k, v["start_index"], v["end_index"], v["owner"])
        return msg

    def parse_all(self, request):
        try:
            with transaction.atomic():
                serializer = TestGuidSerializer(data=request.POST.dict())

                if serializer.is_valid():

                    # get or create guid
                    guid = TestGuid().getFromDictData(request, serializer.validated_data)
                    guid_obj = TestGuid.objects.filter(**guid.get_pk_union_key()).first()
                    if guid_obj:
                        self.sheet_name_map['guid'] = guid_obj.toDic()
                        if request.REQUEST.get('execute_mode') == 'new':
                            raise ICGException('ICG100004', 'Guid is exist, please check your date is right or not,'
                                                            ' or select merge mode to upload.')
                    else:
                        if request.REQUEST.get('execute_mode') == 'merge':
                            raise ICGException('ICG100004', 'Guid hasn\'t exist, please check your date is right or not,'
                                                            'or select new mode to upload')
                        serializer.save()
                        self.sheet_name_map['guid'] = serializer.data
                    # self.sheet_name_map['guid']['exe_type'] = request.REQUEST.get('exe_type', 'Auto')
                    if request.REQUEST.get('execute_mode') == 'merge':
                        self.sheet_name_map['guid']['execute_mode'] = 'merge'
                        if request.REQUEST.get('check_mode_auto'):
                            self.sheet_name_map['guid']['clear_auto'] = True
                        if request.REQUEST.get('check_mode_manual'):
                            self.sheet_name_map['guid']['clear_manual'] = True
                        if request.REQUEST.get('check_mode_summary'):
                            self.sheet_name_map['guid']['clear_summary'] = True
                        if request.REQUEST.get('check_mode_environment'):
                            self.sheet_name_map['guid']['clear_environment'] = True
                        if request.REQUEST.get('check_mode_suite'):
                            self.sheet_name_map['guid']['clear_suite'] = True
                        if request.REQUEST.get('check_mode_bugslist'):
                            self.sheet_name_map['guid']['clear_bugslist'] = True
                    else:
                        self.sheet_name_map['guid']['execute_mode'] = 'new'
                    self.sheet_name_map['guid']['project_id'] = request.REQUEST.get('project_id', '0')
                    #设置guid权限为dashboard
                    self.sheet_name_map['guid']['permission'] = 10
                else:
                    self.sheet_name_map['guid'] = serializer.errors

                for sheet_name in self.sheet_name_map:
                    if sheet_name != 'guid':
                        msg = self.parse(sheet_name)
                        if msg:
                            self.sheet_name_map[sheet_name]['errors'] = msg
                        else:
                            self.check(self.sheet_name_map[sheet_name])
                        # break  # only test extra

                raise RequestDone
        except RequestDone:
            return True
        except Exception as e:
            print traceback.format_exc()
            raise e

    def get_sheet_by_alias(self, sheet_alias):
        sheet_index = None
        sheet_alias = self.trim(sheet_alias)
        for m in self._sheet_name_map:
            if m == sheet_alias:
                sheet_index = self._sheet_name_map[m]['index']
                break
            elif self._sheet_name_map[m]['alias'] == sheet_alias:
                sheet_index = self._sheet_name_map[m]['index']
                break
        # not found
        if sheet_index is None:
            raise Exception("%s not found in %s" % (sheet_alias, ",".join([self._sheet_name_map[m]['alias']
                                                                           for m in self._sheet_name_map]),))

        return self.work_book.sheet_by_index(sheet_index)

    def check_validated_data(self, validated_data, rules, cls):
        flag, msg, sum = False, '', 0
        platform = self.sheet_name_map['guid']['platform']
        if platform not in ['IQStudio', 'SDK']:
            if rules['buildTime']['serializer_cls'] == cls:
                for _index, _i in enumerate(validated_data):
                    #　确认TestEnvironment中有Build Time 信息
                    if _i.get(rules['buildTime']['name']) == rules['buildTime']['required_value']:
                        break
                else:
                    msg, flag = ICGException('ICG100004', r'Build Time is required!').message, True
                return msg, flag

            # 处理求和(Sum)
            # for _sum in rules['sum']:
            #     if _i.get(_sum.get('name')) and _sum.get('serializer_cls') == cls:
            #         try:
            #             sum += float(_i[_sum['name']])
            #             if _index == len(validated_data) - 1:
            #                 if not _sum['range']['min'] < sum < _sum['range']['max']:
            #                     msg, flag = ICGException('ICG100004',_sum['name'] + ' sum must gt ' +
            #                     '%s' % (_sum['range']['min']) + ' or lt '
            #                     + '%s' % (_sum['range']['max'])).message, True
            #         except ValueError:
            #             msg, flag = ICGException('ICG100004', r'weight must be \d+ ').message, True

        return msg, flag

    def update_hsd_url(self, hsd_id):
        HSDLINK = r'https://vthsd.fm.intel.com/hsd/tablet_platform/default.aspx#sighting/default.aspx?sighting_id='
        BZLINK = r'https://bz01p-vied.ir.intel.com/show_bug.cgi?id='
        HSDESLINK = r'https://hsdes.intel.com/home/default.html#article?id='

        hsd_url = 'NA'  # default
        if isinstance(hsd_id, (int, long)):
            hsd_url = HSDLINK + str(hsd_id)

        elif isinstance(hsd_id, str):
            if hsd_id.isdigit():
                hsd_url = HSDLINK + hsd_id
        else:
            if hsd_id.startswith('BZ'):
                hsd_url = BZLINK + hsd_id
            elif hsd_id.startswith('HSD'):
                hsd_url = HSDESLINK + hsd_id.replace('HSD', '')

        return hsd_url

    def update_bug_id(self, bug_id):
        _tmp = re.search('\d+', bug_id)
        if _tmp:
            return _tmp.group()
    #递归查询
    def get_contour_items_all_parent(self, project_id, test_case_id):

        def recursion(obj):
            try:
                obj = ContourItem.objects.get(id=obj.parent_id)
                if obj.parent_id == 0:
                    return
                counter_list.append(obj)
                recursion(obj)
            except ContourItem.DoesNotExist:
                return

        res = ContourItem.objects.filter(project_id=project_id, test_case_id=test_case_id)
        counter_list = []
        if res:
            for i in res:
                recursion(i)

        serializer = CounterItemsSerializer(counter_list, many=True)
        try:
            first_category = serializer.data[-1]['name']
            second_category = serializer.data[-2]['name']
        except:
            first_category, second_category = '',''
        return first_category, second_category


    # #一次查询
    # def get_contour_items_all_parent(self, project_id, test_case_id):
    #
    #     def recursion(parent_id, list):
    #         for i in list:
    #             if i.id == parent_id:
    #                 if i.parent_id == 0:
    #                     return
    #                 id_list.append(i)
    #                 recursion(i.parent_id, list)
    #
    #     qs = ContourItem.objects.filter(project_id=project_id)
    #     id_list = []
    #     data = [i for i in qs]
    #     for i in data:
    #         if i.test_case_id == test_case_id:
    #             recursion(i.parent_id, data)
    #
    #     try:
    #         first_category, second_category = id_list[-1].name, id_list[-2].name
    #     except:
    #         first_category, second_category = '', ''
    #     return first_category, second_category





    # def get_contour_items(self, project_id, test_case_id):
    #
    #     res = ContourItem.objects.filter(project_id=project_id, test_case_id=test_case_id).first()
    #     serializer = CounterItemsSerializer(res)
    #     return serializer.data


class EcrReportParse(object):
    ALLOW_SHEET = ('dashboard', 'summary',)
    OPTIONS_SHEET = ('ecr_status_type', 'review_health_status_type',)
    OPTION_CHECK_LIST = {"ecr_status": {"type": "ecr_status_type"}, "review_health_status": {"type": "review_health_status_type"}}
    SECTION_INDEX = 1
    SECTION_START_MARK = "section_start"
    SECTION_END_MARK = "section_end"
    SECTION_NAME_MAP = {
        "ecr_summary": {"expect_keys": ["highlights", "lowlights", "change_request_query_list", "note"], "type": "vertical", "serializer_cls": "EcrSummarySerializer"},
        "ecr_roadmap": {"dialect": {"ecr_roadmap_snapshot": "snapshot", "delivery_schedule": "deliver"}, "expect_keys": ["snapshot", "deliver", "status", "comment"], "type": "horizontal", "serializer_cls": "EcrRoadmapSerializer"},
        "execution_summary": {"expect_keys": ["new", "committed", "delivered", "zbbed", "withdraw"], "type": "horizontal", "serializer_cls": "EcrExecutionSummarySerializer"},
        "ecr_status": {"alias": "ecr_new_commit_close_status", "dialect": {"id": "link", "project": "project_name"}, "expect_keys": ["link", "ipu", "bu", "project_name", "commit_date", "t_shirt_size", "title", "comment", "type"], "type": "horizontal", "serializer_cls": "EcrStatusSerializer"},
        "review_health_status": {"alias": "active_review", "dialect": {"ecr_id": "link","project": "project_name"}, "expect_keys": ["link","review_id", "ipu", "bu", "project_name", "title", "review_days", "comment", "type"], "type": "horizontal", "serializer_cls": "EcrReviewHealthStatusSerializer"},
        "ecr_trend": {"dialect": {"ecr#_in_new": "in_new", "ecr#_in_pre_scoping":"in_pre_scoping", "ecr#_in_scoping":"in_scoping", "ecr#_in_additional_info":"in_additional","ecr#_in_commit": "in_commit","ecr#_in_close":"in_close"},"expect_keys": ["total_open", "open_per_week", "in_new", "in_pre_scoping", "in_scoping", "in_additional", "in_commit", "in_close", "total_close", "cumulative_close_as_delivery", "cumulative_zbb", "year", "month"], "type": "horizontal", "serializer_cls": "EcrTrendSerializer"},
        "ecr_close_trend": {"alias": "close_ecr_trend", "dialect": {"ecr#_in_new": "in_new", "ecr#_in_pre_scoping":"in_pre_scoping", "ecr#_in_scoping":"in_scoping", "ecr#_in_additional_info":"in_additional","ecr#_in_commit": "in_commit","ecr#_in_close":"in_close"},"expect_keys": ["total_open", "open_per_week", "in_new", "in_pre_scoping", "in_additional", "in_commit", "in_close", "total_close", "cumulative_close_as_delivery","cumulative_zbb", "year", "month"], "type": "horizontal", "serializer_cls": "EcrCloseTrendSerializer"},
        "ecr_tte": {"dialect": {"ecr_number": "number"}, "alias": "tte_(time_to_estimate)_*", "expect_keys": ["first_day", "last_day", "month_point", "ontime", "goal", "number", "within_6_weeks", "year", "month"], "type": "horizontal", "serializer_cls": "EcrTteSerializer"},
        "ecr_review_health_trend": {"dialect": {"weeks_mov_avg(received)": "received_5_weeks_mov_avg","weeks_mov_avg(resolved)": "resolved_5_weeks_mov_avg"}, "expect_keys": ["received", "resolved", "received_5_weeks_mov_avg", "resolved_5_weeks_mov_avg", "year", "month", "ww"], "type": "horizontal", "serializer_cls": "EcrReviewHealthTrendSerializer"},
        "ecr_burndown": {"expect_keys": ["ww", "planned", "actual", "year"], "type": "horizontal", "serializer_cls": "EcrBurndownSerializer"},
        "review_burndown": {"expect_keys": ["ww", "planned", "actual", "year"], "type": "horizontal", "serializer_cls": "EcrReviewBurndownSerializer"},
        "ecr_committed_health_trend":  {"alias":"committed_ecr_health_trend", "dialect": {"planned_delivery": "committed", "actual_delivery": "delivered","weeks_mov_avg(committed)": "committed_5_weeks_mov_avg","weeks_mov_avg(delivered)": "delivered_5_weeks_mov_avg"}, "expect_keys": ["committed", "delivered", "committed_5_weeks_mov_avg", "delivered_5_weeks_mov_avg", "ww"  , "year", "month"], "type": "horizontal", "serializer_cls": "EcrCommittedHealthTrendSerializer"},
        "ecr_effort_staff": {"alias": "ecr_effort_sum_of_cumulative_delivered_per_staff_team_(01_17_to_now)", "expect_keys": ["team", "man_week", "year", "month",], "type": "horizontal", "serializer_cls": "EcrEffortStaffSerializer"},
        "ecr_effort_ipu": {"alias": "ecr_effort_sum_of_cumulative_delivered_per_ipu_(01_17_to_now)", "expect_keys": ["ipu", "android","linux", "windows", "chrome", "year", "month", ], "type": "horizontal", "serializer_cls": "EcrEffortIPUSerializer"},
        "ecr_t_shirt_size": {"alias": "#_of_active_committed_ecr_per_t_shirt_size", "expect_keys": ["total_effort", "ecr_number", "year", "month", ],"type": "horizontal", "serializer_cls": "EcrTShirtSizeSerializer"},
        "ecr_active_committed": {"alias": "#_of_team_support_for_active_committed_ecr", "expect_keys": ["team_number", "ecr_number", "year", "month", ], "type": "horizontal","serializer_cls": "EcrActiveCommittedSerializer"},
    }

    def __init__(self, *args, **kwargs):
        self.file_stream = kwargs.get('file_stream')

        if self.file_stream is None:
            raise Exception('excel data not found!')
        else:
            self._work_book = None
            self._sheet_name_map = OrderedDict()

        self._work_book = openpyxl.load_workbook(filename=BytesIO(self.file_stream.read()), data_only=True)
        for index, sheet in enumerate(self._work_book.worksheets):
            if sheet.sheet_state != 'hidden':
                self._sheet_name_map[self.trim(sheet.title)] = {"alias": sheet.title, "index": index, "sections": OrderedDict()}

    @property
    def work_book(self):
        return self._work_book

    @property
    def sheet_name_map(self):
        return self._sheet_name_map

    @classmethod
    def trim(cls, value):
        return re.sub(r'[-:;"\',\\/\s\n]+', '_', str(value).lower())

    def format(self, lookup_keys, expect_keys, optional_keys, dialect):
        _lookup_keys = {}
        for index, lookup in enumerate(lookup_keys):
            lookup = self.trim(lookup.strip())
            for d, k in dialect.items():
                if d == lookup:
                    _lookup_keys[index] = k
                    break
            else:
                if lookup:
                    _lookup_keys[index] = lookup
        miss_key = set(expect_keys) - set(_lookup_keys.values())
        if miss_key:
            raise ICGException('ICG000001', ','.join(miss_key))
        return _lookup_keys

    def index(self, section, col_name):
        index = -1
        for k, v in section.items():
            if v == col_name:
                index = k
                break
        return index

    def check(self, sheet_name, section_map, request):
        try:
            with transaction.atomic():
                sheet = self.get_sheet_by_alias(section_map['alias'])

                for k, v in section_map['sections'].items():
                    # merge keys list
                    data = []
                    self.sheet_name_map['guid'] = {}
                    keys_list = self.format(v['key_list'], v['expect_keys'], v.get('optional_keys', []), v.get('dialect', {}))

                    if sheet_name == 'dashboard':
                        serializer_cls = eval(v['serializer_cls'])
                        model_cls = serializer_cls.Meta.model
                        # insert or update guid
                        for row in sheet.iter_rows(min_row=v['start_index'] + 1, max_row=v['end_index']):
                            cols = [cell.value for cell in row]
                            raw = dict(zip(keys_list.values(), cols))
                            if raw.get('ww') and str(raw.get('ww')).lower() == 's':
                                raw['ww'] = 0
                            if hasattr(serializer_cls, 'pop_null_number'):
                                for pop_num in getattr(serializer_cls, 'pop_null_number', []):
                                    if pop_num in raw and raw[pop_num] is None:
                                        raw.pop(pop_num)
                            guid_obj = EcrGuid.objects.filter(year=raw['year'], month=raw['month']).first()
                            if guid_obj:
                                # update
                                raw['guid'] = guid_obj.id
                                qs = model_cls.objects.filter(guid=guid_obj.id)
                                for lp in eval(v['serializer_cls']).lookup_fields:
                                    qs = eval("qs.filter("+lp+"='"+str(raw.get(lp)).replace("'", "\\\'")+"')")
                                if qs:
                                    qs.delete()
                            else:
                                guid_dict = {'year': raw['year'], 'month': raw['month']}
                                serializer = EcrGuidSerializer(data=guid_dict)
                                if serializer.is_valid():
                                    serializer.save()
                                    raw['guid'] = serializer.data['id']
                                else:
                                    self.sheet_name_map['guid']['errors'] = serializer.errors
                            data.append(raw)
                        serializer = serializer_cls(data=data, many=True)
                        if serializer.is_valid():
                            v['validated_data'] = serializer.data
                            serializer.save()
                        else:
                            v['errors'] = serializer.errors

                    else:
                        # get or create guid
                        guid = EcrGuid().getFromDictData(request, request.data)
                        guid_obj = EcrGuid.objects.filter(**guid.get_pk_union_key()).first()
                        if guid_obj:
                            # update guid description
                            if guid.description and guid.description != guid_obj.description:
                                guid_obj.description = guid.description
                                guid_obj.save()
                            self.sheet_name_map['guid'] = guid_obj.toDic()
                            if v['serializer_cls'] in EcrReportSerializer.get_summary_group():
                                eval(v['serializer_cls'].replace('Serializer', '')).objects.filter(guid=guid_obj.id).delete()
                        else:
                            guid_dict = guid.toDic()
                            guid_dict.pop('id')
                            serializer = EcrGuidSerializer(data=guid_dict)
                            if serializer.is_valid():
                                serializer.save()
                                self.sheet_name_map['guid'] = serializer.data
                            else:
                                self.sheet_name_map['guid']['errors'] = serializer.errors

                        # detail part
                        if v["type"] == 'horizontal':
                            for row in sheet.iter_rows(min_row=v['start_index']+1, max_row=v['end_index']):
                                cols = []
                                for cell in row:
                                    if cell.hyperlink:
                                        hyperlink = cell.hyperlink.target
                                        if cell.hyperlink.location:
                                            hyperlink += "#" + cell.hyperlink.location
                                        cols.append('<a href="%s" target="_blank">%s</a>' % (hyperlink, cell.value))
                                    else:
                                        cols.append(cell.value)

                                # cols = [r.value for r in row]
                                raw = dict(zip(keys_list.values(), cols))
                                raw['guid'] = self.sheet_name_map['guid']['id']
                                for option, verify_lst in self.OPTION_CHECK_LIST.get(k, {}).items():
                                    if raw[option] not in getattr(self, verify_lst, []):
                                        raise ICGException('ICG100010', option + " " + raw[option] + " not in "
                                                           + str(getattr(self, verify_lst, [])))
                                data.append(raw)
                        elif v["type"] == 'vertical':
                            raw = {'guid': self.sheet_name_map['guid']['id']}
                            for row_num, key in keys_list.items():
                                cell = sheet.cell(row=row_num+v['start_index'], column=self.SECTION_INDEX)
                                if cell.coordinate in sheet.merged_cells:
                                    for merge in sheet.merged_cell_ranges:
                                        if merge.startswith(cell.coordinate + ":"):
                                            merge_split = merge.split(':')
                                            merge_start = openpyxl.utils.coordinate_from_string(merge_split[0])[1]
                                            merge_end = openpyxl.utils.coordinate_from_string(merge_split[1])[1]
                                            raw[key] = ''
                                            for merge_index in range(merge_start, merge_end+1):
                                                c = sheet.cell(row=merge_index, column=self.SECTION_INDEX+1)
                                                if c.hyperlink:
                                                    hyperlink = c.hyperlink.target
                                                    if c.hyperlink.location:
                                                        hyperlink += "#" + c.hyperlink.location
                                                    raw[key] += '<a href="%s" target="_blank">%s</a><br/>' % (hyperlink, c.value)
                                                else:
                                                    raw[key] += '' if c.value is None else c.value
                                            break
                                else:
                                    raw[key] = sheet.cell(row=row_num+v['start_index'], column=self.SECTION_INDEX+1).value
                            data.append(raw)

                        serializer = eval(v['serializer_cls'])(data=data, many=True)
                        if serializer.is_valid():
                            v['validated_data'] = serializer.data
                            serializer.save()
                        else:
                            v['errors'] = serializer.errors
            # raise RequestDone
        except RequestDone:
            return True
        except ICGException as e:
            v['errors'] = e.error_detail
        except Exception as e:
            v['errors'] = str(e)
            print traceback.format_exc()

    def get_sheet_by_alias(self, sheet_alias):
        return self.work_book.get_sheet_by_name(sheet_alias)

    def parse(self, sheet_name):
        sheet = self.get_sheet_by_alias(sheet_name)
        section = msg = None
        sections = OrderedDict()

        # get section start&end index
        for row in range(sheet.min_row, sheet.max_row + self.SECTION_INDEX):
            section_mark = str(sheet.cell(row=row, column=self.SECTION_INDEX).value).strip()
            if self.trim(section_mark) == self.SECTION_START_MARK:
                # section mark followed section name
                section = str(sheet.cell(row=row, column=self.SECTION_INDEX+1).value).strip()
                sections.setdefault(self.trim(section), {})
                # mark data start index
                sections[self.trim(section)]["start_index"] = row

            if self.trim(section_mark) == self.SECTION_END_MARK:
                if section == str(sheet.cell(row=row, column=self.SECTION_INDEX+1).value).strip() and section:
                    sections.setdefault(self.trim(section), {})
                    # mark data end index
                    sections[self.trim(section)]["end_index"] = row
                else:
                    msg = str(sheet.cell(row=row, column=self.SECTION_INDEX+1).value).strip() + " section start flag not found"

        for k, v in sections.items():
            if v.get("start_index") is None:
                msg = k + " section start flag not found"
            if v.get("end_index") is None:
                msg = k + " section end flag not found"

            if v.get("start_index") is not None and v.get("end_index") is not None:
                start_index = v['start_index'] + 1
                end_index = v['end_index'] - 1
                for _s in self.SECTION_NAME_MAP:
                    if self.SECTION_NAME_MAP[_s].get('alias', _s) == k:
                        _SECTION_NAME_MAP = self.SECTION_NAME_MAP[_s]
                        break
                else:
                    return False
                key_list = []
                if _SECTION_NAME_MAP.get("type") == 'horizontal':
                    for row in sheet.iter_rows(min_row=start_index, max_row=start_index):
                        for r in row:
                            key_list.append(str(r.value).strip(' ').strip('\n'))
                elif _SECTION_NAME_MAP.get("type") == 'vertical':
                    for row in sheet.iter_rows(min_row=start_index, max_row=end_index):
                        key_list.append(str(row[self.SECTION_INDEX - 1].value).strip(' ').strip('\n'))
                else:  # no_head_data
                    key_list = _SECTION_NAME_MAP["expect_keys"]

                # set data start_index & end index & key list
                self._sheet_name_map[self.trim(sheet.title)]["sections"][k] = {
                    "key_list": key_list,
                    "start_index": start_index,
                    "end_index": end_index,
                }
                self._sheet_name_map[self.trim(sheet.title)]["sections"][k].update(
                    _SECTION_NAME_MAP
                )

            if msg:
                break

        return msg

    def parse_all(self, request):
        # get option
        for option in self.OPTIONS_SHEET:
            sheet = self.get_sheet_by_alias(self.sheet_name_map[option]['alias'])
            option_list = []
            for row in sheet.iter_rows():
                option_list.append(row[0].value)
            setattr(self, option, option_list)
            cache.set('ERC_REPORT_'+option, option_list, 365 * 24 * 60 * 60)
        # check format
        for sheet_name in self.sheet_name_map:
            if sheet_name in self.ALLOW_SHEET:
                msg = self.parse(self.sheet_name_map[sheet_name]['alias'])
                if msg:
                    self.sheet_name_map[sheet_name]['errors'] = msg
                else:
                    self.check(sheet_name, self.sheet_name_map[sheet_name], request)

            else:
                self.sheet_name_map.pop(sheet_name)  # remove other sheet


if __name__ == "__main__":
    pass
    # print xlrd.book.XL_CELL_EMPTY, xlrd.book.XL_CELL_NUMBER
    # print type(cell.value), type(UNICODE_LITERAL('')
    # print cell.ctype, xlrd.book.XL_CELL_TEXT

    # validationReportParse = ValidationReportParse(**{"file_path": './IPU4-IoTG-Linux - PIT-Yocto_IVI_WW53.2.xlsx'})
    # validationReportParse.save(temporary=True)
    #
    # validationReportParse = ValidationReportParse.restore(memory_key="temporary_file")

    # validationReportHandle = ValidationReportParse(file_path='./IPU4-IoTG-Linux - PIT-Yocto_IVI_WW53.2.xlsx')
    # print validationReportHandle.parse(sheet_name='Extra')['extra']




